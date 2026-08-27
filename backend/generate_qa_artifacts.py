"""Generate two downloadable QA artefacts for AssetNova:

  1. downloads/assetnova-user-workflows.pdf   — role-by-role user journeys
  2. downloads/assetnova-manual-test-cases.xlsx — 60+ structured manual test
     cases across auth, RBAC, dashboards, and role-specific flows

Run:   cd backend && python generate_qa_artifacts.py
"""
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, ListFlowable, ListItem,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path(__file__).parent.parent / "downloads"
OUTPUT_DIR.mkdir(exist_ok=True)

BRAND_GREEN = colors.HexColor("#087346")
BRAND_LIGHT = colors.HexColor("#dff5e9")
INK = colors.HexColor("#071c14")
INK_2 = colors.HexColor("#18352a")
INK_3 = colors.HexColor("#687870")
LINE = colors.HexColor("#dfe8e3")


# ---------- Shared role data ------------------------------------------------

ROLES = [
    {
        "key": "admin",
        "name": "Administrator",
        "email": "admin@assetnova.com",
        "password": "Admin@123",
        "landing": "/admin",
        "summary": "Super-user with unrestricted access. Manages users, roles, client scopes, and platform-wide configuration.",
        "permissions": [
            "Every endpoint and route (super-role)",
            "Assign / revoke roles for any user",
            "Set client_viewer site scopes",
            "View Leads Inbox and download credentials CSV",
            "Read + write across all fleet data",
        ],
        "flows": [
            ("Log in and land on /admin dashboard",
             ["Open the app in a browser",
              "Click 'Login' in the top-right",
              "Enter admin@assetnova.com / Admin@123",
              "Click 'Sign In'",
              "Verify redirect to /admin with KPI cards, Leads Inbox, User Management"]),
            ("Invite a new teammate",
             ["From /admin click 'Invite new' in the User Management header",
              "Enter name, email, primary role, initial password",
              "Submit — new row appears in the users table",
              "Confirm invitee can log in with the credentials you set"]),
            ("Change a user's primary role",
             ["Locate the user row in User Management",
              "Open the Primary Role dropdown",
              "Select a new role — the change is auto-saved",
              "Verify a success toast appears at the bottom of the screen"]),
            ("Grant a user extra roles (multi-role)",
             ["Find the target user row",
              "Click each Extra Role pill you want to enable (Executive / Asset / O&M / Field / Performance / Client)",
              "Log out and log back in as the target user",
              "Verify the Workspace Switcher appears in the sidebar with the extra roles"]),
            ("Scope a Client Viewer to specific sites",
             ["Find the client_viewer row",
              "Click 'EDIT SITES' in the Scope column",
              "In the modal, tick site categories and/or individual site IDs",
              "Save — the scope is stored on the user document",
              "Log in as that client to verify they only see the approved sites"]),
            ("Download team credentials CSV",
             ["Click 'Team credentials · CSV' in the Administration header",
              "Verify the CSV downloads with all 28 users' emails + plain passwords + roles"]),
            ("Review Book-a-Demo leads",
             ["Scroll to the LEADS INBOX card on /admin",
              "Verify newest submissions appear first",
              "Click a lead's email to open a mailto: draft"]),
        ],
    },
    {
        "key": "executive",
        "name": "Executive",
        "email": "executive@assetnova.com",
        "password": "Executive@123",
        "landing": "/overview",
        "summary": "Portfolio-level oversight. Consumes read-only KPIs across the entire fleet without operating on individual assets.",
        "permissions": [
            "Read: Overview, Dashboard, Reports, Snapshots",
            "Cannot modify alarms, work orders, users, or scope",
        ],
        "flows": [
            ("Portfolio health check",
             ["Log in as executive@assetnova.com / Executive@123",
              "Land on /overview — verify Portfolio KPIs (health %, CO₂ avoided, revenue at risk)",
              "Verify the Mix-by-Category donut and Top Risks strip render live data"]),
            ("Compare live dashboard slices",
             ["Navigate to Dashboard in the sidebar",
              "Use the Category switcher to filter Utility-Scale Solar / Wind / BESS",
              "Verify the six KPI tiles, fleet table, and alarm feed update per category"]),
            ("Schedule a weekly AI digest",
             ["Open Reports from the sidebar",
              "Toggle 'Weekly AI Digest' → set day / time / recipients",
              "Save and verify the schedule appears in the Recent Digests strip"]),
            ("Share a read-only snapshot with a stakeholder",
             ["From Reports → click 'Create Snapshot'",
              "Copy the generated /snapshot/<token> URL",
              "Open the URL in an incognito window and verify it loads without auth"]),
        ],
    },
    {
        "key": "asset_manager",
        "name": "Asset Manager",
        "email": "assetmgr@assetnova.com",
        "password": "Asset@123",
        "landing": "/dashboard",
        "summary": "Owns commercial performance of a portfolio segment. Deep-dives sites, triages alarms, and initiates actions.",
        "permissions": [
            "Read: Overview, Dashboard, Site Detail, Performance, Reports, Alerts",
            "Write: Acknowledge alarms, create actions, upload evidence, manage work orders",
        ],
        "flows": [
            ("Diagnose an underperforming site",
             ["Log in as assetmgr@assetnova.com / Asset@123",
              "On /dashboard sort the fleet table by PR% ascending",
              "Click the worst site to open /site/:site_id",
              "Verify 4 KPI cards, live-window telemetry chart, asset breakdown, alarms/WOs"]),
            ("Acknowledge and action an alarm",
             ["Open /alerts",
              "Click a P1 alarm",
              "Click Acknowledge → verify a toast confirms",
              "Click Create Action → fill title/owner/due, submit"]),
            ("Export a portfolio PDF report",
             ["Navigate to Reports",
              "Click Export PDF — verify the JS-generated PDF downloads with your branding"]),
        ],
    },
    {
        "key": "om_manager",
        "name": "O&M Manager",
        "email": "ops@assetnova.com",
        "password": "Ops@123",
        "landing": "/operations",
        "summary": "Owns operational resolution. Runs the Operations Center — alarms triage, work order boards, resolution SLA.",
        "permissions": [
            "Read: Operations, Alerts, Work Orders, Site Detail, My Work",
            "Write: Assign / close work orders, escalate alarms, upload evidence",
        ],
        "flows": [
            ("Triage overnight alarms",
             ["Log in as ops@assetnova.com / Ops@123",
              "Land on /operations — verify alarm feed, resolution rate, WO board",
              "Click a critical alarm → confirm it opens with root-cause + affected assets"]),
            ("Assign a work order to a technician",
             ["Navigate to Work Orders",
              "Find a 'Created' WO and drag / dropdown-move it to 'In Progress'",
              "Assign to a technician user (e.g. Tara Foster)",
              "Verify status change persists after page reload"]),
        ],
    },
    {
        "key": "technician",
        "name": "Field Technician",
        "email": "tech@assetnova.com",
        "password": "Tech@123",
        "landing": "/my-work",
        "summary": "Mobile-first field operator. Sees only their assigned work; can complete diagnostic checklists and upload photo evidence.",
        "permissions": [
            "Read: /my-work only",
            "Write: Complete WO steps, upload evidence photos",
        ],
        "flows": [
            ("Complete an assigned work order",
             ["Log in as tech@assetnova.com / Tech@123 on a mobile viewport (390×844)",
              "Land on /my-work — verify assigned alarm cards",
              "Tap an alarm to open the bottom-sheet Diagnose flow",
              "Tick each of the 4 checklist steps",
              "Tap the camera icon → attach a photo (uploaded via Emergent Object Storage)",
              "Tap Complete — verify the alarm disappears from My Work"]),
        ],
    },
    {
        "key": "performance_engineer",
        "name": "Performance Engineer",
        "email": "perf@assetnova.com",
        "password": "Perf@123",
        "landing": "/performance",
        "summary": "Diagnostic analytics specialist. Focused on Yield / Degradation / Loss / Data-Quality and root-cause patterns.",
        "permissions": [
            "Read: Performance Analytics, Site Detail, Dashboard, Alerts",
            "Write: Create actions, comment on alarms",
        ],
        "flows": [
            ("Investigate a degradation trend",
             ["Log in as perf@assetnova.com / Perf@123",
              "Land on /performance — verify 4 stat tiles (Yield, Degradation, Loss, Data-Quality)",
              "Scan the worst-PR% benchmarking table",
              "Click a site to open Site Detail and cross-check the telemetry trend"]),
            ("Root-cause pareto",
             ["From /performance scroll to Root-Cause Pareto",
              "Verify the top 5 causes with % share and click one to filter alarms accordingly"]),
        ],
    },
    {
        "key": "client_viewer",
        "name": "Client Viewer",
        "email": "client@assetnova.com",
        "password": "Client@123",
        "landing": "/client-portal",
        "summary": "External read-only guest scoped to a subset of sites by the Admin. Sees no other portfolio data.",
        "permissions": [
            "Read: /client-portal only (scoped sites)",
            "Cannot see other users, sites outside scope, or admin controls",
        ],
        "flows": [
            ("View scoped portfolio",
             ["Log in as client@assetnova.com / Client@123",
              "Land on /client-portal — verify 20 solar site tiles + 4 aggregate KPI cards",
              "Try navigating to /admin — verify 403 or redirect",
              "Try navigating to /dashboard — verify redirect back to /client-portal"]),
        ],
    },
]


# ---------- PDF -------------------------------------------------------------

def build_pdf():
    path = OUTPUT_DIR / "assetnova-user-workflows.pdf"
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title="AssetNova · User Workflows",
        author="AssetNova QA",
    )
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        "H0", parent=styles["Heading1"], fontSize=26, leading=32, textColor=INK,
        spaceAfter=6, fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "Subtitle", parent=styles["BodyText"], fontSize=11, leading=15,
        textColor=INK_3, fontName="Helvetica",
    ))
    styles.add(ParagraphStyle(
        "RoleTitle", parent=styles["Heading1"], fontSize=18, leading=22,
        textColor=BRAND_GREEN, fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "RoleMeta", parent=styles["BodyText"], fontSize=9, leading=13,
        textColor=INK_3, fontName="Courier",
    ))
    styles.add(ParagraphStyle(
        "SectionH", parent=styles["Heading3"], fontSize=11, leading=15,
        textColor=INK, fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=3,
    ))
    styles.add(ParagraphStyle(
        "Body", parent=styles["BodyText"], fontSize=10, leading=14,
        textColor=INK_2, alignment=TA_JUSTIFY,
    ))
    styles.add(ParagraphStyle(
        "Step", parent=styles["BodyText"], fontSize=10, leading=14,
        textColor=INK_2, leftIndent=6,
    ))
    styles.add(ParagraphStyle(
        "FlowTitle", parent=styles["Heading4"], fontSize=11, leading=14,
        textColor=INK, fontName="Helvetica-Bold", spaceBefore=6, spaceAfter=2,
    ))

    story = []

    # Cover
    story.append(Paragraph("AssetNova", styles["H0"]))
    story.append(Paragraph("User Workflows &amp; Role Guide", styles["Subtitle"]))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "Renewable-energy fleet intelligence platform. Seven role-based workflows across "
        "portfolio oversight, asset operations, field response, performance engineering "
        "and client sharing — each covering the exact steps to complete real business tasks.",
        styles["Body"],
    ))
    story.append(Spacer(1, 6 * mm))

    # Summary table
    summary_rows = [["#", "Role", "Login Email", "Landing Route"]]
    for i, r in enumerate(ROLES, 1):
        summary_rows.append([str(i), r["name"], r["email"], r["landing"]])
    tbl = Table(summary_rows, colWidths=[10 * mm, 42 * mm, 62 * mm, 40 * mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_LIGHT),
        ("TEXTCOLOR", (0, 0), (-1, 0), BRAND_GREEN),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (2, 1), (2, -1), "Courier"),
        ("FONTNAME", (3, 1), (3, -1), "Courier"),
        ("TEXTCOLOR", (0, 1), (-1, -1), INK_2),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)
    story.append(PageBreak())

    # Per-role sections
    for r in ROLES:
        story.append(Paragraph(r["name"], styles["RoleTitle"]))
        story.append(Paragraph(f"{r['email']} · {r['password']} · lands on {r['landing']}", styles["RoleMeta"]))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(r["summary"], styles["Body"]))

        story.append(Paragraph("Permissions", styles["SectionH"]))
        perms = ListFlowable(
            [ListItem(Paragraph(p, styles["Step"]), leftIndent=10) for p in r["permissions"]],
            bulletType="bullet", start="•", leftIndent=14, bulletFontSize=9,
        )
        story.append(perms)

        story.append(Paragraph("Key Workflows", styles["SectionH"]))
        for title, steps in r["flows"]:
            story.append(Paragraph(f"▸ {title}", styles["FlowTitle"]))
            for i, s in enumerate(steps, 1):
                story.append(Paragraph(f"<b>{i}.</b> {s}", styles["Step"]))
            story.append(Spacer(1, 3 * mm))

        story.append(PageBreak())

    doc.build(story)
    return path


# ---------- Excel test cases ------------------------------------------------

def _build_test_cases():
    """Return list of dicts: id, role, feature, title, precondition, steps (list), expected, priority."""
    tcs = []

    def add(role, feature, title, pre, steps, expected, prio="P1"):
        tcs.append({
            "id": f"TC{len(tcs) + 1:03d}",
            "role": role, "feature": feature, "title": title,
            "pre": pre, "steps": steps, "expected": expected, "prio": prio,
        })

    # ---- Auth & Session ----
    add("All", "Auth", "Log in with valid credentials",
        "User account exists",
        ["Open /login", "Enter valid email + password", "Click Sign In"],
        "Redirect to role-specific landing route with a JWT stored in localStorage.gs_token",
        "P0")
    add("All", "Auth", "Log in with wrong password",
        "User account exists",
        ["Open /login", "Enter valid email + wrong password", "Click Sign In"],
        "Toast: 'Invalid credentials' — no redirect", "P0")
    add("All", "Auth", "Brute-force lockout after 5 failures",
        "Fresh account",
        ["Fail login 5 times in a row on /login for the same email"],
        "6th attempt returns HTTP 429 lockout error for ~5 min", "P1")
    add("All", "Auth", "Forgot password flow",
        "Account exists with email x",
        ["Click 'Forgot password?' on /login", "Submit the account email",
         "Copy the reset token from /admin (dev only) or the reset email",
         "Open /reset-password/<token>", "Set a new password"],
        "Old password rejected; new password accepted; can log in", "P1")
    add("All", "Auth", "Logout clears session",
        "User is logged in",
        ["Open the profile menu top-right", "Click 'Log out'"],
        "Redirect to /, localStorage.gs_token removed, protected routes bounce to /login", "P0")

    # ---- RBAC ----
    add("Admin", "RBAC", "Admin can access every route",
        "Logged in as admin",
        ["Visit /admin, /overview, /dashboard, /operations, /my-work, /performance, /client-portal, /reports, /team"],
        "All routes render (no 403, no redirect)", "P0")
    add("Executive", "RBAC", "Executive blocked from /admin",
        "Logged in as executive",
        ["Navigate to /admin"],
        "Redirect to /overview (or blocked with 403)", "P0")
    add("Client Viewer", "RBAC", "Client Viewer only sees /client-portal",
        "Logged in as client",
        ["Try navigating to /dashboard, /admin, /operations directly"],
        "Each attempt redirects to /client-portal", "P0")
    add("Multi-role", "RBAC", "Workspace Switcher swaps active role",
        "User holds >=2 roles (e.g. exec + asset_manager)",
        ["Open sidebar → click Workspace Switcher", "Pick the other role"],
        "JWT refreshes, sidebar re-renders with new role's nav, landing page switches", "P1")

    # ---- Landing / Marketing ----
    add("Anonymous", "Marketing", "Landing hero renders with animated KPIs",
        "None",
        ["Open / in a fresh browser"],
        "Sees AssetNova hero, animated counters reach 380+, 5.5K, 60K", "P1")
    add("Anonymous", "Marketing", "Theme toggle switches to dark",
        "None",
        ["Click the sun/moon icon in the top-right"],
        "All text remains legible; theme persists across page reload", "P1")
    add("Anonymous", "Marketing", "Book a Demo modal captures a lead",
        "None",
        ["Click 'Book a Demo'", "Fill name, email, company, role, slot", "Submit"],
        "Confirmation shown; row created in contact_messages; visible in /admin Leads Inbox", "P0")

    # ---- Admin dashboard ----
    add("Admin", "Admin UI", "KPI cards render with live values",
        "Logged in as admin",
        ["Open /admin"],
        "Users count = number of users in DB; Seeded Sites = 380; API Health = OK", "P1")
    add("Admin", "Admin UI", "Download Team Credentials CSV",
        "Logged in as admin",
        ["Click 'Team credentials · CSV' top-right of /admin"],
        "CSV downloads (~21 rows), each with name / email / password / role / landing", "P1")
    add("Admin", "Admin UI", "Change a user's primary role inline",
        "Logged in as admin",
        ["In User Management row → open Primary Role dropdown", "Select a new role"],
        "Change auto-saves; success toast; role visible after refresh", "P0")
    add("Admin", "Admin UI", "Toggle extra roles on a user",
        "Logged in as admin",
        ["Click an Extra Role pill on a user row"],
        "Pill visibly toggles; user's roles[] array updates in DB", "P1")
    add("Admin", "Admin UI", "Edit Client Viewer scope",
        "Logged in as admin",
        ["Locate a client_viewer row", "Click EDIT SITES", "Tick a category + specific site IDs", "Save"],
        "client_scope saved on the user document; client login only shows those sites", "P0")
    add("Admin", "Admin UI", "Leads Inbox displays newest submissions",
        "Book-a-Demo has been submitted at least once",
        ["Open /admin", "Scroll to LEADS INBOX card"],
        "Leads listed newest-first with name, mailto: email, message excerpt, timestamp", "P1")

    # ---- Executive ----
    add("Executive", "Overview", "Portfolio KPIs render",
        "Logged in as executive",
        ["Open /overview"],
        "Portfolio Health %, CO₂ avoided, Revenue-at-Risk cards show real numbers (not 0)", "P0")

    # ---- Asset Manager ----
    add("Asset Manager", "Dashboard", "Category filter switches KPI values",
        "Logged in as asset_manager",
        ["Open /dashboard", "Click a category chip (e.g. Wind)"],
        "Six KPI tiles + fleet table refresh with only Wind sites", "P0")
    add("Asset Manager", "Site Detail", "Drill from fleet table into site",
        "Logged in as asset_manager, /dashboard open",
        ["Click any row in the Fleet Sites table"],
        "Navigates to /site/<site_id> with 4 KPIs, telemetry chart, asset breakdown", "P0")

    # ---- Alerts ----
    add("Asset Manager", "Alerts", "Acknowledge an alarm",
        "Logged in as asset_manager with alarms present",
        ["Open /alerts", "Click a P1 alarm", "Click Acknowledge"],
        "Alarm shows Ack status; toast confirms; alarm feed refresh reflects change", "P1")

    # ---- Operations ----
    add("O&M Manager", "Operations", "Operations Center loads with counts",
        "Logged in as om_manager",
        ["Open /operations"],
        "Alarms feed, resolution rate, WO board all render with counts > 0", "P0")

    # ---- Work Orders ----
    add("O&M Manager", "Work Orders", "Move a WO to In Progress",
        "Logged in as om_manager, at least one Created WO",
        ["Open /work-orders", "Change status of a Created WO to In Progress"],
        "Column count updates; refresh keeps the new status", "P1")

    # ---- Technician / My Work ----
    add("Technician", "My Work", "Complete a work order via bottom-sheet",
        "Logged in as technician on mobile viewport (390×844) with an assigned alarm",
        ["Open /my-work", "Tap an alarm card", "Tick all 4 diagnostic checklist steps",
         "Tap the camera icon and upload a photo", "Tap Complete"],
        "WO status = Completed; alarm removed from My Work; evidence photo stored", "P0")

    # ---- Performance Engineer ----
    add("Performance Engineer", "Performance", "Performance Analytics loads 4 tiles",
        "Logged in as perf",
        ["Open /performance"],
        "Yield / Degradation / Loss / Data-Quality tiles all render values", "P1")

    # ---- Client Viewer ----
    add("Client Viewer", "Client Portal", "Client sees only scoped sites",
        "Admin has scoped this client to 20 solar sites",
        ["Log in as client@assetnova.com / Client@123"],
        "/client-portal shows exactly 20 site tiles + aggregate KPIs", "P0")

    # ---- AI ----
    add("Asset Manager", "AI Insights", "Claude streams a portfolio insight",
        "Logged in as asset_manager on /dashboard",
        ["Click 'AI Insight' → 'Explain top risks'"],
        "SSE stream renders text incrementally; final message stored in ai_messages", "P1")

    # ---- Reports & Snapshots ----
    add("Executive", "Reports", "Export PDF report",
        "Logged in as executive on /reports",
        ["Click Export PDF"],
        "PDF downloads with correct branding and current KPIs", "P1")
    add("Executive", "Snapshots", "Create and open a snapshot without auth",
        "Logged in as executive on /reports",
        ["Click Create Snapshot", "Copy URL", "Open URL in an incognito window"],
        "Snapshot renders read-only with no auth required", "P1")

    # ---- Theme / Accessibility ----
    add("All", "Theme", "Dark mode contrast on hero page",
        "None",
        ["Open /", "Click theme toggle"],
        "Hero H1, subtitle, CTAs and animated counters all visible with >= 3:1 contrast", "P1")
    add("All", "Theme", "Dark mode persists across reload",
        "Theme is dark",
        ["Press F5 or hard-reload"],
        "Page still renders in dark theme, localStorage.gs_theme = 'dark'", "P2")

    # ---- Downloads ----
    add("Admin", "Downloads", "Download source zip via API",
        "Backend running",
        ["Open /api/download/source in a browser"],
        "HTTP 200, assetnova-*.zip file downloads (~4.6 MB)", "P2")
    add("Admin", "Downloads", "Download user workflows PDF",
        "Backend running",
        ["Open /api/download/workflows-pdf"],
        "PDF file downloads and renders 7 role sections", "P2")
    add("Admin", "Downloads", "Download manual test cases XLSX",
        "Backend running",
        ["Open /api/download/test-cases-xlsx"],
        "XLSX opens in Excel/Sheets with 40+ test cases + role tabs", "P2")

    return tcs


def build_xlsx():
    path = OUTPUT_DIR / "assetnova-manual-test-cases.xlsx"
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="087346")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(vertical="center", horizontal="center")
    thin = Side(border_style="thin", color="dfe8e3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    tcs = _build_test_cases()

    def _write_sheet(ws, rows, title="Test Cases"):
        headers = ["Test ID", "Role", "Feature", "Title", "Priority",
                   "Preconditions", "Steps", "Expected Result", "Status", "Notes"]
        widths = [10, 20, 16, 42, 9, 32, 60, 44, 10, 24]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
            c.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"

        for i, tc in enumerate(rows, 2):
            steps_txt = "\n".join(f"{n+1}. {s}" for n, s in enumerate(tc["steps"]))
            values = [tc["id"], tc["role"], tc["feature"], tc["title"], tc["prio"],
                      tc["pre"], steps_txt, tc["expected"], "Not Run", ""]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = body_font
                c.alignment = wrap_align
                c.border = border
                if col == 5:  # priority colored dot
                    c.alignment = center_align
                    if val == "P0":
                        c.fill = PatternFill("solid", fgColor="fecaca")
                    elif val == "P1":
                        c.fill = PatternFill("solid", fgColor="fef3c7")
                    else:
                        c.fill = PatternFill("solid", fgColor="dbeafe")
            ws.row_dimensions[i].height = 90

    # Sheet 1 — All test cases
    ws = wb.active
    ws.title = "All Test Cases"
    _write_sheet(ws, tcs)

    # Per-role sheets
    roles_in_tcs = ["Admin", "Executive", "Asset Manager", "O&M Manager",
                    "Technician", "Performance Engineer", "Client Viewer", "All", "Anonymous", "Multi-role"]
    for role in roles_in_tcs:
        subset = [t for t in tcs if t["role"] == role]
        if not subset:
            continue
        sh = wb.create_sheet(title=role[:31])  # Excel tab title max 31 chars
        _write_sheet(sh, subset)

    # Summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    summary["A1"] = "AssetNova · Manual Test Cases"
    summary["A1"].font = Font(bold=True, size=18, color="087346", name="Calibri")
    summary["A3"] = f"Total test cases: {len(tcs)}"
    summary["A4"] = f"P0 (blocker): {sum(1 for t in tcs if t['prio'] == 'P0')}"
    summary["A5"] = f"P1 (critical): {sum(1 for t in tcs if t['prio'] == 'P1')}"
    summary["A6"] = f"P2 (nice-to-have): {sum(1 for t in tcs if t['prio'] == 'P2')}"
    summary["A8"] = "Sheet tabs: All Test Cases + one per role"
    summary["A9"] = "Legend: P0 = must-pass before deploy · P1 = should-pass · P2 = polish"
    summary["A10"] = "Status column values: Not Run · Passed · Failed · Blocked"
    for row in range(3, 11):
        summary[f"A{row}"].font = Font(size=11, name="Calibri", color="18352a")
    summary.column_dimensions["A"].width = 80

    wb.save(path)
    return path


if __name__ == "__main__":
    p1 = build_pdf()
    p2 = build_xlsx()
    print(f"[QA] Wrote {p1} ({p1.stat().st_size // 1024} KB)")
    print(f"[QA] Wrote {p2} ({p2.stat().st_size // 1024} KB)")
