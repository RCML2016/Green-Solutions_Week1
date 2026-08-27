"""
Idempotently seeds the AssetNova dataset (Sites, Assets, Telemetry, Weather,
Performance, Alarms, Work_Orders) into MongoDB from the Excel workbook shipped in
/app/backend/data/green_solutions_sample_data.xlsx.

Run automatically on backend startup if the `fleet_sites` collection is empty.
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

DATA_FILE = Path(__file__).parent / "data" / "green_solutions_sample_data.xlsx"

SHEET_COLLECTION_MAP = {
    "Sites": "fleet_sites",
    "Assets": "fleet_assets",
    "Telemetry": "fleet_telemetry",
    "Weather": "fleet_weather",
    "Performance": "fleet_performance",
    "Alarms": "fleet_alarms",
    "Work_Orders": "fleet_work_orders",
}

# Per-collection indexes
INDEXES = {
    "fleet_sites": [("site_id", 1), ("site_type", 1), ("state", 1)],
    "fleet_assets": [("site_id", 1), ("asset_id", 1), ("asset_type", 1)],
    "fleet_telemetry": [("site_id", 1), ("asset_id", 1), ("timestamp", -1)],
    "fleet_weather": [("site_id", 1), ("timestamp", -1)],
    "fleet_performance": [("site_id", 1), ("date", -1)],
    "fleet_alarms": [("site_id", 1), ("timestamp", -1), ("severity", 1), ("status", 1)],
    "fleet_work_orders": [("site_id", 1), ("status", 1), ("alarm_id", 1)],
}


def _row_to_doc(header: List[str], row: tuple) -> Dict[str, Any]:
    doc: Dict[str, Any] = {}
    for key, val in zip(header, row):
        if isinstance(val, datetime):
            # store as ISO string so JSON serialisation is trivial
            if val.tzinfo is None:
                val = val.replace(tzinfo=timezone.utc)
            doc[key] = val.isoformat()
        else:
            doc[key] = val
    return doc


async def seed_if_empty(db) -> Dict[str, int]:
    """Insert dataset into Mongo iff `fleet_sites` is empty. Returns per-collection counts."""
    existing_sites = await db.fleet_sites.count_documents({})
    if existing_sites > 0:
        return {"skipped": existing_sites}

    if not DATA_FILE.exists():
        logging.warning("[SEED] Data file not found at %s — skipping seed.", DATA_FILE)
        return {"missing_file": True}

    logging.info("[SEED] Loading AssetNova dataset from %s …", DATA_FILE)
    wb = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

    result: Dict[str, int] = {}
    for sheet_name, collection_name in SHEET_COLLECTION_MAP.items():
        if sheet_name not in wb.sheetnames:
            logging.warning("[SEED] Sheet %s missing in workbook", sheet_name)
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            continue

        batch: List[Dict[str, Any]] = []
        total = 0
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            batch.append(_row_to_doc(header, row))
            if len(batch) >= 2000:
                await db[collection_name].insert_many(batch, ordered=False)
                total += len(batch)
                batch = []
        if batch:
            await db[collection_name].insert_many(batch, ordered=False)
            total += len(batch)
        result[collection_name] = total
        logging.info("[SEED] %s → %s (%d docs)", sheet_name, collection_name, total)

    # Indexes
    for coll_name, keys in INDEXES.items():
        try:
            await db[coll_name].create_index(keys)
        except Exception as e:  # noqa: BLE001
            logging.warning("[SEED] index on %s failed: %s", coll_name, e)

    wb.close()
    logging.info("[SEED] Done. %s", result)
    return result
