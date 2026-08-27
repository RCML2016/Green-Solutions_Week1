"""Iteration 18 — QA artefact download endpoints (workflows PDF, test-cases XLSX, credentials CSV regression)."""
import os

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    return s


class TestDownloadArtifacts:
    def test_workflows_pdf(self, client):
        r = client.get(f"{BASE_URL}/api/download/workflows-pdf", timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert r.headers.get("content-type", "").startswith("application/pdf"), r.headers
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower()
        assert "assetnova-user-workflows.pdf" in cd
        assert r.content[:5] == b"%PDF-", r.content[:20]
        assert len(r.content) > 5000, len(r.content)

    def test_test_cases_xlsx(self, client):
        r = client.get(f"{BASE_URL}/api/download/test-cases-xlsx", timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert r.headers.get("content-type", "") == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), r.headers.get("content-type")
        cd = r.headers.get("content-disposition", "")
        assert "assetnova-manual-test-cases.xlsx" in cd
        assert r.content[:2] == b"PK", r.content[:20]
        assert len(r.content) > 10000, len(r.content)

    def test_xlsx_content_structure(self, client):
        """Validate the workbook actually parses and has structured test cases."""
        import io

        from openpyxl import load_workbook

        r = client.get(f"{BASE_URL}/api/download/test-cases-xlsx", timeout=60)
        wb = load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) >= 2, wb.sheetnames
        # find total data rows across sheets
        total_rows = 0
        headers_seen = None
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hdr = [str(c).strip().lower() if c else "" for c in rows[0]]
            if "test id" in hdr or "test_id" in hdr:
                headers_seen = hdr
                total_rows += len([row for row in rows[1:] if any(row)])
        assert headers_seen is not None, f"No sheet with Test ID header: {wb.sheetnames}"
        for col in ["role", "priority", "steps", "expected result"]:
            assert any(col in h for h in headers_seen), (col, headers_seen)
        assert total_rows >= 40, f"Only {total_rows} test-case rows found"

    def test_credentials_csv_regression(self, client):
        r = client.get(f"{BASE_URL}/api/download/team-credentials", timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert "text/csv" in r.headers.get("content-type", "")
        body = r.content.decode("utf-8", "replace")
        assert "@assetnova.com" in body
        assert len(body.splitlines()) >= 10

    def test_source_zip_still_available(self, client):
        r = client.get(f"{BASE_URL}/api/download/source", timeout=120, stream=True)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/zip")
        r.close()


class TestAdminAuthRegression:
    def test_admin_login(self, client):
        r = client.post(f"{BASE_URL}/api/auth/login", json={
            "email": "admin@assetnova.com", "password": "Admin@123"})
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        token = data.get("access_token") or data.get("token")
        assert token, data
        assert data.get("user", {}).get("role") == "admin"
